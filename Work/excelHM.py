import openpyxl
import pandas as pd
import os 
from googletrans import Translator

fn = "DataFromHandM.xlsx"



book = openpyxl.load_workbook(fn)
sheet = book.active
# translator = Translator()

malomerit = 'По отзывам наших покупателей, данная модель идет в размер!'
v_razmer = 'По отзывам наших покупателей, данная модель маломерит!'
bolshemerit = 'По отзывам наших покупателей, данная модель большемерит!'
capitals = ['IDs','Title','Color','Description','Material','Bolshemerit','Premium','Size','Concept','Photos','']

sheet.insert_cols(1, amount=len(capitals))

#Formula in column 
def formula(column):
    for i in range(2, len(column_I_data) + 1): 
        sheet.cell(row=i, column=column).value = f"=VLOOKUP(A{i}, M:V, {column}, FALSE)"


for i in range(1, sheet.max_row + 1):
    value = sheet.cell(row=i, column=13).value
    sheet.cell(row=i, column=1).value = value

# Copy data from column I to a list
column_I_data = [sheet.cell(row=i, column=13).value for i in range(2, sheet.max_row + 1)]
# Sort the data in ascending order
column_I_data.sort()
# Remove duplicates
column_I_data = list(dict.fromkeys(column_I_data))


# Clear existing data in column A
for i in range(1, sheet.max_row + 1):
    sheet.cell(row=i, column=1).value = None

# Write sorted and deduplicated data to column A
for i, value in enumerate(column_I_data, start=1):
    sheet.cell(row=i, column=1).value = value


# Fill the formula down to the last row of data in the columns
formula(2)
formula(3)
formula(4)
formula(5)


# Fill the column with formulas
#Photos
for i in range(2, len(column_I_data) + 1):
    sheet.cell(row=i, column=10).value = f'=SUBSTITUTE(VLOOKUP(A{i}, M:V, 10, FALSE), "//lp2.hm.com/hmgoepprod?set=", "https://lp2.hm.com/hmgoepprod?set=")'
#Premium
for i in range(2, len(column_I_data) + 1):
    sheet.cell(row=i, column=7).value = f'=IF(ISNUMBER(SEARCH("H&M Premium Selection", VLOOKUP(A{i}, M:V, 7, FALSE))), VLOOKUP(A{i}, M:V, 7, FALSE), "")'
#Size
for i in range(2, len(column_I_data) + 1):
    sheet.cell(row=i, column=8).value = f'=IF(ISNUMBER(SEARCH("Rozmiar", VLOOKUP(A{i}, M:V, 8, FALSE))), VLOOKUP(A{i}, M:V, 8, FALSE), "")'
#Concept
for i in range(2, len(column_I_data) + 1):
    sheet.cell(row=i, column=9).value = f'=IF(ISNUMBER(SEARCH("Dział", VLOOKUP(A{i}, M:V, 9, FALSE))), VLOOKUP(A{i}, M:V, 9, FALSE), "")'
#Bolshemerit
for i in range(2, len(column_I_data) + 1):
    sheet.cell(row=i, column=6).value = f'=IFERROR(CHOOSE(MATCH(VALUE(MID(VLOOKUP(A{i}, M:V, 6, FALSE),16,2)),' + '{0,41,60,101}, 1), ' + f'"{malomerit}", "{v_razmer}", "{bolshemerit}"), "")'




# Fill the names in the columns
for i, capital in enumerate(capitals, start=1):
    sheet.cell(row=1, column=i, value=capital)

# Save the workbook
book.save("HM_updated.xlsx")

